#!/usr/bin/env python3
import io
import numpy as np
import cv2 as cv
import time
import numpy as np
#import matplotlib
import rospy
import math
from cv_bridge       import CvBridge
from sensor_msgs.msg import Image
import cv2 as cv
from std_msgs.msg import String
import threading
import openpyxl
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error

def canny(image):
    gray_img = cv.cvtColor(image, cv.COLOR_RGB2GRAY)
    blur_img = cv.GaussianBlur(gray_img, (1, 1), 0)
    canny_img = cv.Canny(blur_img, 50, 200, None, 3)
    return canny_img

def line_detection (masked_img):
    lines_list = []
    lines = cv.HoughLinesP(
        masked_img,  # Input edge image
        1,  # Distance resolution in pixels
        np.pi / 180,  # Angle resolution in radians
        threshold=50,  # Min number of votes for valid line
        minLineLength=50,  # Min allowed length of line
        maxLineGap=3 # Max allowed gap between line for joining them
    )
    return lines
def region_of_interest(image):
    height = 288
    width = 512
    polygons = np.array([[(0, 250), (512, 250), (256,180)]])
    mask = np.zeros_like(image)
    cv.rectangle( mask, (0, 250), (width, height), 255, -1)
    cv.fillPoly(mask, polygons, 255)
    masked_image = cv.bitwise_and(image, mask)
    cv.imshow("masked", masked_image)
    return masked_image
def dist( x1, x2 , y1, y2 ):
    d = (x1 - x2)**2 + (y1 - y2)**2
    return d


# 
# def angle(x1,x2,y1,y2, offset):
#     if y1 == y2 :
#         return 90;
#     else:
#         a_1 = (x1-x2)/(y1-y2)
#         t = math.atan(a_1)*57.32 - offset
#         return t

def angle (lines):
    ang_arr = np.zeros(1)
    for i in range(len(lines)) :
        points = lines[i]
        x1, y1, x2, y2 = points[0]
        if y1 == y2 :
            t = 90
        else:
            a_1 = (x1-x2)/(y1-y2)
            t = math.atan(a_1)*57.32
        if i == 0 :
            ang_arr[0] = t
        else:
            ang_arr = np.append(ang_arr, t)
    return  ang_arr

def get_hue(image):
    # Chuyển đổi không gian màu từ BGR sang HSV
    hsv_img = cv.cvtColor(image, cv.COLOR_BGR2HSV)
    gray_img = hsv_img[:, :, 0]
    #cv.imshow("hue frame", gray_img)
    gray_img = cv.equalizeHist(gray_img)
    gray_img = cv.GaussianBlur(gray_img, (7, 5), 0)
    canny_img = cv.Canny(gray_img, 50, 200, None, 3)
    return canny_img
def decode(V ,off):
    v1 = V + off*V
    v2 = V - off*V

    return v1, v2
def discrete_pid(Kp, Ki, Kd, dt, max_output, min_output):
    # Khởi tạo các biến
    error_sum = 0
    last_error = 0

    def pid(error):
        nonlocal error_sum, last_error

        # Tính toán tổng lỗi tích phân
        error_sum += error * dt

        # Tính toán đạo hàm của lỗi
        d_error = (error - last_error) / dt
        last_error = error

        # Tính toán giá trị điều khiển mới
        output = Kp * error + Ki * error_sum + Kd * d_error

        # Giới hạn giá trị điều khiển trong khoảng cho phép
        if output > max_output:
            output = max_output
        elif output < min_output:
            output = min_output

        return output
def discrete_pid(Kp, Ki, Kd, dt, max_output, min_output):
    # Khởi tạo các biến
    error_sum = 0
    last_error = 0

    def pid(error):
        nonlocal error_sum, last_error

        # Tính toán tổng lỗi tích phân
        error_sum += error * dt

        # Tính toán đạo hàm của lỗi
        d_error = (error - last_error) / dt
        last_error = error

        # Tính toán giá trị điều khiển mới
        output = Kp * error + Ki * error_sum + Kd * d_error

        # Giới hạn giá trị điều khiển trong khoảng cho phép
        if output > max_output:
            output = max_output
        elif output < min_output:
            output = min_output

        return output

    return pid

def equal_hsv(image):
    hsv_img = cv.cvtColor(image, cv.COLOR_BGR2HSV)
    hsv_img[:, :, 2] = cv.equalizeHist(hsv_img[:, :, 2])
    image = cv.cvtColor(hsv_img, cv.COLOR_HSV2BGR)
    return image
def ransac(X, Y):
    best_model = None
    best_error = np.inf
    for i in range(40):  # Số lần lấy mẫu
        sample_idx = np.random.choice(len(X), size=2, replace=False)
        X_sample = X[sample_idx]
        Y_sample = Y[sample_idx]
        model = LinearRegression().fit(X_sample.reshape(-1, 1), Y_sample)
        Y_pred = model.predict(X.reshape(-1, 1))
        error = mean_squared_error(Y, Y_pred)
        if error < best_error:
            best_model = model
            best_error = error
    return best_model

def error1(x1, x2, y1, y2):
    e = 0
    if y1 != y2:
        # x = ay + b
        a = (x2 - x1) / (y2 - y1)
        b = x1 - a * y1
        e = a*180 + b
    else:
        e = -1

    return int(256 - e)

angle_confirm = 0
data = openpyxl.load_workbook('/home/pi/Desktop/get_line.xlsx')
sheet1 = data['Sheet1']
def ransac_confirm(img):
    my_array = np.array([[0, 0]], dtype=np.uint8)
    global sheet1, angle_confirm
    for j in range(1, 9):
        for i in range(1, 237):
            x = sheet1.cell(row=i, column=2 * j - 1).value
            y = sheet1.cell(row=i, column=2 * j).value
            if ((j == 6 or j == 12) and i == 62) or ((j == 7 or j == 11) and i == 39) or (
                    (j == 8 or j == 10) and i == 18) or ((j == 4 or j == 14 ) and i == 179 or (
                    (j == 5 or j == 13) and i == 100)):
                break
            elif img[y][x] == 255:
                # cv.line(img, (256, 287), (x, y), 255, 1)
                # cv.line(Original_frame, (256, 287), (x, y), (0, 255, 0), 1)
                my_array = np.append(my_array,[[x, y]], axis= 0)
                break
    my_array = my_array[1:]

    X = my_array[:, 0]
    Y = my_array[:, 1]

    model = ransac(X, Y)
    angle_confirm = 90 - model.coef_[0]*57.3


# Khởi tạo hàm PID
pid = discrete_pid(Kp=0.0005, Ki=0, Kd= 0, dt=0.2, max_output= 1 , min_output= -1)

temp = 1 # Line or lane
lane = 0
lane_Thr = 30 # switch delay

bridge = CvBridge()
Original_frame = None ;

def callback(msg):
        global angel_confirm, error
        Original_frame =bridge.imgmsg_to_cv2(msg)
        frame = canny(Original_frame)
        start = time.time()
        thread = threading.Thread(target=ransac_confirm, args=(frame,))
        thread.start()

        lines_list = []
        masked_img = region_of_interest(frame)
        lines = line_detection(masked_img)
        max_d = 0
        d_arr = np.zeros(1)

        if lines is not None:
            for points in lines:
                # Extracted points nested in the list
                x1, y1, x2, y2 = points[0]
                #cv.line(Original_frame, (x1, y1), (x2, y2), (255, 0, 0), 2)
                d = dist(x1, x2, y1, y2)
                if d_arr[0] == 0:
                    d_arr[0] = d
                else:
                    d_arr = np.append(d_arr, d)


            # error = angle(x1_m, x2_m, y1_m, y2_m, 66)
            # ransac_confirm(frame)
            thread.join()
            sorted_indices = np.argsort(d_arr)[::-1] # long line sort

            ang_arr = angle(lines)

            near_arr = np.abs(ang_arr - angle_confirm)

            #sort_arr = (sorted_indices + 1)*near_arr

            sort_arr1 = - d_arr * 0.002+ near_arr
            
            #idx = np.argsort(np.abs(ang_arr - angle_confirm)) # near angle_comfirm sort
            idx1 = np.argsort(sort_arr1)
            #print(ang_arr[idx1[0]])
            #error  = ang_arr[idx1[0]]
            #rospy.loginfo("error = " + str(error))
            x1, y1, x2, y2 = lines[idx1[0]][0]
            error = error1(x1, x2, y1, y2)
            rospy.loginfo("error =" + str(error))
            cv.line(Original_frame, (x1, y1), (x2, y2), (0, 255, 0), 2)



        end = time.time()
        cv.putText(Original_frame, 'fps = ' + str(1 / (end - start)), (10, 10), cv.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0),
                    2)
        #cv.putText(Original_frame, 'angle = ' + str(error), (10, 10), cv.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0),
        #            2)
        cv.circle(Original_frame, (- error + 256, 180), 2, (0, 0, 255), -1)
        #cv.imshow("edge_detection", frame)
        cv.imshow("line_detectiion", Original_frame)
        
def decition_callback(msg):
    global Decition, Decition_pre
    Decition_pre = Decition
    Decition = str(msg.data)

rospy.init_node("line_detection_node", anonymous=True)        
rospy.Subscriber("/automobile/image_raw",Image,callback)
pub = rospy.Publisher('command', String, queue_size=10)
rospy.Subscriber('decition',String, decition_callback)
error = 0
Decition = 'S'
Decition_pre = 'S'
while not rospy.is_shutdown():
    control = pid(error)
    #rospy.loginfo("control =" + str(control))
    V1, V2 = decode(60, control)
    if Decition == 'R' and Decition_pre == 'S':
        command = "!V0;0#"
        pub.publish(command)
        time.sleep(1)
        command = "!V-45;45#"
        pub.publish(command)
        time.sleep(1.5)
        command = "!V0;0#"
        pub.publish(command)
        time.sleep(0.5)
    
    command = "!V" +str(int(V1)) + ";" + str(int(V2)) + "#"
    #command = "!V0;0#"
    #rospy.loginfo("command ="+ command)
    pub.publish(command)
    time.sleep(0.2)
    if cv.waitKey(1) & 0xFF == ord('q'):
        break
